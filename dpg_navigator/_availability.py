"""Dependency probes for optional preview backends.

Each ``*_available()`` is safe at import time (failed imports stay
``None``). HTML/Chrome probes delegate to ``_html.py``.
"""

from __future__ import annotations

from ._optional import OptionalCallable, OptionalModule, as_optional

# PDF
_pdfium: OptionalModule | None
try:
    import pypdfium2 as _pdfium_mod  # type: ignore[import-untyped]

    _pdfium = as_optional(_pdfium_mod)
except Exception:
    _pdfium = None
_np: OptionalModule | None
try:
    import numpy as _numpy_mod

    _np = as_optional(_numpy_mod)
except Exception:
    _np = None
_PILImage: OptionalModule | None
try:
    from PIL import Image as _PILImage_mod

    _PILImage = as_optional(_PILImage_mod)
except Exception:
    _PILImage = None


def pdf_available() -> bool:
    """Return True if all PDF preview dependencies are installed."""
    return _pdfium is not None and _np is not None and _PILImage is not None


# HTML & Chrome — canonical probes live in _html.py (html2image + numpy +
# Pillow, and a cached Chrome binary lookup). Imported lazily so this module
# stays GUI-free at import time.
def html_available() -> bool:
    """Return True if html2image, numpy, and Pillow are installed."""
    from ._html import html_available as _html_available

    return _html_available()


def chrome_available() -> bool:
    """Return True if HTML packages are present and a Chrome binary is found.

    Delegates to ``dpg_navigator._html.chrome_available`` (cached; honors
    ``DPG_CHROME_BIN`` / ``CHROME_BIN`` / ``CHROME_PATH``).
    """
    from ._html import chrome_available as _chrome_available

    return _chrome_available()


# Word
_DocxDocument: OptionalCallable | None
try:
    from docx import Document as _DocxDocument_cls

    _DocxDocument = as_optional(_DocxDocument_cls)
except Exception:
    _DocxDocument = None
_mammoth: OptionalModule | None
try:
    import mammoth as _mammoth_mod  # type: ignore[import-untyped]

    _mammoth = as_optional(_mammoth_mod)
except Exception:
    _mammoth = None


def word_available() -> bool:
    """Return True if Word text-extraction dependencies are installed."""
    return _DocxDocument is not None


def mammoth_available() -> bool:
    """Return True if mammoth + html2image Word preview is available."""
    return _mammoth is not None and html_available()


# PowerPoint
_Presentation: OptionalCallable | None
try:
    from pptx import Presentation as _Presentation_cls

    _Presentation = as_optional(_Presentation_cls)
except Exception:
    _Presentation = None


def pptx_available() -> bool:
    """Return True if PowerPoint preview dependencies are installed."""
    return _Presentation is not None


# Excel
_load_workbook: OptionalCallable | None
try:
    from openpyxl import load_workbook as _load_workbook_fn  # type: ignore[import-untyped]

    _load_workbook = as_optional(_load_workbook_fn)
except Exception:
    _load_workbook = None


def excel_available() -> bool:
    """Return True if Excel (.xlsx) preview dependencies are installed."""
    return _load_workbook is not None


# Markdown
_markdown: OptionalModule | None
try:
    import markdown as _markdown_mod  # type: ignore[import-untyped]

    _markdown = as_optional(_markdown_mod)
except Exception:
    _markdown = None


def markdown_available() -> bool:
    """Return True if rendered Markdown preview is available."""
    return _markdown is not None and html_available()


# Pygments — installed for PreviewKind.CODE routing; highlighting is not rendered.
_pygments: OptionalModule | None
try:
    import pygments as _pygments_mod  # type: ignore[import-untyped]

    _pygments = as_optional(_pygments_mod)
except Exception:
    _pygments = None


def pygments_available() -> bool:
    """Return True if Pygments is installed.

    Gates ``PreviewKind.CODE`` routing. Source files still render as
    monospace text — Pygments is not used for highlighting.
    """
    return _pygments is not None


# 7z
_py7zr: OptionalModule | None
try:
    import py7zr as _py7zr_mod

    _py7zr = as_optional(_py7zr_mod)
except Exception:
    _py7zr = None


def seven_zip_available() -> bool:
    """Return True if py7zr dependencies are installed for .7z support."""
    return _py7zr is not None


def py7zr_available() -> bool:
    """Return True if py7zr dependencies are installed for .7z support."""
    return seven_zip_available()
