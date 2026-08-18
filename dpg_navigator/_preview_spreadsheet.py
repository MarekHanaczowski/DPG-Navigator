"""Excel metadata loading for the preview panel.

Parses workbook sheets into table-ready rows without depending on DearPyGui.
"""

from __future__ import annotations

# MIT licensed
from dataclasses import dataclass
from typing import Any, Callable

from ._optional import OptionalCallable, as_optional

_load_workbook: OptionalCallable | None
try:
    from openpyxl import load_workbook as _load_workbook_fn  # type: ignore[import-untyped]

    _load_workbook = as_optional(_load_workbook_fn)
except Exception:  # optional backend absent or incompatible (e.g. old Python)
    _load_workbook = None


class ExcelPreviewError(Exception):
    """Excel workbook data could not be loaded."""


@dataclass(frozen=True)
class SpreadsheetTable:
    """Table-ready worksheet data."""

    headers: list[str]
    rows: list[list[str]]
    status: str
    sheetnames: list[str]
    sheet_name: str


def excel_available() -> bool:
    """Return True when Excel workbook loading is available."""
    return _load_workbook is not None


def load_excel_table(
    path: str,
    *,
    sheet_name: str | None,
    max_rows: int,
    max_cols: int,
    workbook_loader: Callable[..., Any] | None = _load_workbook,
) -> SpreadsheetTable:
    """Load one worksheet into bounded table-ready data."""
    if workbook_loader is None:
        raise ExcelPreviewError("openpyxl is not installed")

    try:
        workbook = workbook_loader(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelPreviewError(str(exc)) from exc

    try:
        sheetnames = list(workbook.sheetnames)
        worksheet = None
        if sheet_name and sheet_name in sheetnames:
            worksheet = workbook[sheet_name]
        elif sheetnames:
            worksheet = workbook.active if workbook.active is not None else workbook[sheetnames[0]]

        if worksheet is None:
            return SpreadsheetTable([], [], "No sheets found", sheetnames, "")

        selected_sheet = worksheet.title
        # Header + displayed rows + one sentinel to detect truncation. Do not
        # walk the rest of the sheet (wide/tall workbooks stall the preview).
        scan_limit = max_rows + 2
        fetched: list[list[str]] = []
        widest_row = 0

        for row in worksheet.iter_rows(max_row=scan_limit, values_only=True):
            widest_row = max(widest_row, len(row))
            fetched.append([str(cell) if cell is not None else "" for cell in row[:max_cols]])
    except Exception as exc:
        raise ExcelPreviewError(str(exc)) from exc
    finally:
        workbook.close()

    if not fetched:
        return SpreadsheetTable(
            [],
            [],
            f"Sheet: {selected_sheet} | Empty sheet",
            sheetnames,
            selected_sheet,
        )

    headers = fetched[0]
    rows = fetched[1:]
    row_capped = len(fetched) >= scan_limit
    if row_capped:
        rows = rows[:max_rows]
        total_label = f"{max_rows}+"
    else:
        total_label = str(len(rows))
    display_cols = min(widest_row, max_cols)

    while len(headers) < display_cols:
        headers.append(f"Col{len(headers) + 1}")

    parts = [
        f"Sheet: {selected_sheet}",
        f"{total_label} rows \u00d7 {widest_row} cols",
    ]
    truncated = []
    if row_capped:
        truncated.append(f"first {len(rows)} rows")
    if display_cols < widest_row:
        truncated.append(f"first {display_cols} cols")
    if truncated:
        parts.append(f"(showing {', '.join(truncated)})")

    return SpreadsheetTable(headers, rows, " | ".join(parts), sheetnames, selected_sheet)
