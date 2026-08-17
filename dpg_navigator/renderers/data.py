"""Data preview renderer for CSV, Excel, SQLite, and XML."""

from __future__ import annotations  # PEP 604/585 in signatures need this on py3.8/3.9

import logging

import dearpygui.dearpygui as dpg
from defusedxml import minidom as _minidom  # type: ignore[import-untyped]

try:
    import openpyxl
except ImportError:
    openpyxl = None
try:
    import sqlite3
except ImportError:
    sqlite3 = None

from typing import Callable

from .._availability import _load_workbook
from .._filesystem import DirectoryLister
from .._preview_registry import PreviewCapabilities
from .._preview_spreadsheet import ExcelPreviewError, load_excel_table
from .._preview_sqlite import SQLitePreviewError, load_sqlite_table
from .._preview_table import CsvPreviewError, parse_csv_table
from .._types import FileEntry
from ._base import BaseRenderer, PreviewContext, TableRenderMixin

_log = logging.getLogger(__name__)


class DataRenderer(TableRenderMixin, BaseRenderer):
    """Render CSV, spreadsheet, SQLite, and XML data in native DPG tables."""

    # _STATUS_HEIGHT is provided by TableRenderMixin.
    _TABLE_MAX_ROWS: int = 200
    _TABLE_MAX_COLS: int = 50
    _TEXT_PREVIEW_MAX_SIZE: int = 256 * 1024
    _text_offset: int = 0

    def __init__(self, load_text_content_cb: Callable[[str, int], tuple[str | None, bool]]):
        self._load_text_content = load_text_content_cb
        self._current_entry = None
        self._ctx = None

    def render(self, entry: FileEntry, ctx: PreviewContext) -> None:
        """Render a supported data file into the preview panel."""
        self._ctx = ctx
        self._current_entry = entry
        ext = entry.ext
        if ext in (".csv", ".tsv", ".psv"):
            self._render_csv_preview(entry)
        elif ext in (".xlsx", ".xlsm"):
            self._render_excel_preview(entry)
        elif ext in (".sqlite", ".sqlite3", ".db"):
            self._render_sqlite_preview(entry)
        elif ext == ".xml":
            self._render_xml_preview(entry)
        else:
            ctx.show_error("Unsupported data format", f"{ext} is not supported")

    def clear(self) -> None:
        """Release the current data preview context."""
        self._current_entry = None
        self._ctx = None

    def _render_binary_warning(self, entry: FileEntry) -> None:
        """Show a 'binary file' notice in the panel (no text preview available)."""
        if self._ctx is None or self._ctx.panel_id is None:
            return
        if self._ctx.temp_font is not None:
            if dpg.does_item_exist(self._ctx.temp_font):
                dpg.delete_item(self._ctx.temp_font)
            self._ctx.temp_font = None
        dpg.delete_item(self._ctx.panel_id, children_only=True)
        tex_tag = f"_preview_tex_{self._ctx.config_tag}"
        if dpg.does_item_exist(tex_tag):
            dpg.delete_item(tex_tag)
        dpg.add_text(
            f"Binary file: {entry.name}",
            color=[128, 128, 128],
            parent=self._ctx.panel_id,
        )
        dpg.add_text(
            "(No text preview available)",
            color=[100, 100, 100],
            parent=self._ctx.panel_id,
        )

    def _render_text_preview(self, entry: FileEntry) -> None:
        """Read a text file and display its contents (fallback for data paths)."""
        if self._ctx is None or self._ctx.panel_id is None:
            return

        text, is_bin = self._load_text_content(entry.full_path, self._text_offset)
        if is_bin:
            self._render_binary_warning(entry)
            return
        if text is None:
            self.clear()
            return

        if not text.strip():
            text = "(No text content or only whitespace in this fragment)"

        self._ctx.image_cache = None
        dpg.delete_item(self._ctx.panel_id, children_only=True)
        tex_tag = f"_preview_tex_{self._ctx.config_tag}"
        if dpg.does_item_exist(tex_tag):
            dpg.delete_item(tex_tag)

        if entry.size_bytes is not None and entry.size_bytes > self._TEXT_PREVIEW_MAX_SIZE:
            self._render_text_navigation(entry)
        else:
            dpg.add_text(
                entry.name,
                color=[180, 180, 255],
                parent=self._ctx.panel_id,
            )
        dpg.add_separator(parent=self._ctx.panel_id)
        with dpg.child_window(parent=self._ctx.panel_id, height=-1, width=-1):
            dpg.add_text(text, wrap=0)

    def _render_text_navigation(self, entry: FileEntry) -> None:
        """Show a static byte-range label for an oversized text/XML fallback.

        DataRenderer has no repaint callback for paging, so it reports the
        displayed byte range instead of creating non-functional controls.
        """
        if self._ctx is None or self._ctx.panel_id is None:
            return
        if entry.size_bytes is None:
            dpg.add_text(entry.name, color=[180, 180, 255], parent=self._ctx.panel_id)
            return
        size_bytes = entry.size_bytes
        mb = 1024 * 1024
        start_mb = self._text_offset / mb
        end_mb = min(
            (self._text_offset + self._TEXT_PREVIEW_MAX_SIZE) / mb,
            size_bytes / mb,
        )
        total_mb = size_bytes / mb
        with dpg.group(horizontal=True, parent=self._ctx.panel_id):
            dpg.add_text(entry.name, color=[180, 180, 255])
            dpg.add_spacer(width=4)
            dpg.add_text(
                f"{start_mb:.2f}-{end_mb:.2f} of {total_mb:.2f} MB",
                color=[200, 200, 200],
            )

    def _render_csv_preview(self, entry: FileEntry) -> None:
        """Parse a CSV/TSV file and display as a native DPG table."""
        if self._ctx is None or self._ctx.panel_id is None:
            return

        text, is_bin = self._load_text_content(entry.full_path)
        if is_bin:
            self._render_binary_warning(entry)
            return
        if text is None:
            self.clear()
            return

        try:
            table = parse_csv_table(
                text,
                entry.name,
                max_rows=self._TABLE_MAX_ROWS,
                max_cols=self._TABLE_MAX_COLS,
            )
        except CsvPreviewError:
            # If CSV parsing fails, fallback to plain text
            self._render_text_preview(entry)
            return

        # If the underlying file exceeds the text page size, indicate partial read.
        header_name = entry.name
        if entry.size_bytes is not None and entry.size_bytes > self._TEXT_PREVIEW_MAX_SIZE:
            limit_str = DirectoryLister.format_size(self._TEXT_PREVIEW_MAX_SIZE)
            size_str = DirectoryLister.format_size(entry.size_bytes)
            header_name = f"{entry.name} (Partial: first {limit_str} of {size_str})"

        self._render_table_widget(
            header_name,
            table.headers,
            table.rows,
            table.status,
        )

    def _render_excel_preview(self, entry: FileEntry, sheet_name_to_load: str | None = None) -> None:
        """Parse an Excel .xlsx file and display as a native DPG table."""
        if self._ctx is None or self._ctx.panel_id is None:
            return

        try:
            table = load_excel_table(
                entry.full_path,
                sheet_name=sheet_name_to_load,
                max_rows=self._TABLE_MAX_ROWS,
                max_cols=self._TABLE_MAX_COLS,
                workbook_loader=_load_workbook,
            )
        except ExcelPreviewError:
            self.clear()
            return

        def _build_excel_ui():
            if len(table.sheetnames) > 1:
                with dpg.group(horizontal=True, parent=self._ctx.panel_id):
                    dpg.add_text("Sheet:", color=[180, 180, 180])

                    def on_sheet_changed(sender, app_data, user_data):
                        self._render_excel_preview(entry, sheet_name_to_load=app_data)

                    dpg.add_combo(
                        items=table.sheetnames, default_value=table.sheet_name, width=-1, callback=on_sheet_changed
                    )

        self._render_table_widget(entry.name, table.headers, table.rows, table.status, ui_builder=_build_excel_ui)

    def _render_xml_preview(self, entry: FileEntry) -> None:
        """Parse XML with ``defusedxml`` and display pretty-printed contents."""
        if self._ctx is None:
            panel_id = getattr(self, "_panel_id", None)
            if panel_id is None:
                return
            self._ctx = PreviewContext(
                panel_id=panel_id,
                table_wrapper=0,
                config_tag=getattr(self, "_config_tag", ""),
                capabilities=PreviewCapabilities(),
            )
        if self._ctx.panel_id is None:
            return

        raw_text, is_bin = self._load_text_content(entry.full_path, self._text_offset)
        if is_bin:
            self._render_binary_warning(entry)
            return
        if raw_text is None:
            self.clear()
            return

        # Format XML
        try:
            parsed = _minidom.parseString(raw_text)
            formatted_text = parsed.toprettyxml(indent="    ")
            # Minidom often adds awkward blank lines, so we clean them up
            text = "\n".join(line for line in formatted_text.splitlines() if line.strip())
        except Exception:
            # If parsing fails (invalid XML), fallback to raw text
            text = raw_text

        if not text:
            text = "(No identifiable XML or text content in this fragment)"

        self._ctx.image_cache = None
        dpg.delete_item(self._ctx.panel_id, children_only=True)
        tex_tag = f"_preview_tex_{self._ctx.config_tag}"
        if dpg.does_item_exist(tex_tag):
            dpg.delete_item(tex_tag)

        # Label info
        if entry.size_bytes is not None and entry.size_bytes > self._TEXT_PREVIEW_MAX_SIZE:
            self._render_text_navigation(entry)
        else:
            dpg.add_text(
                entry.name,
                color=[180, 180, 255],
                parent=self._ctx.panel_id,
            )

        dpg.add_separator(parent=self._ctx.panel_id)
        with dpg.child_window(parent=self._ctx.panel_id, height=-1, width=-1):
            dpg.add_text(text, wrap=0)

    def _render_sqlite_preview(self, entry: FileEntry, table_name_to_load: str | None = None) -> None:
        """Parse a SQLite database file and display a table's contents."""
        if self._ctx is None or self._ctx.panel_id is None:
            return

        try:
            table = load_sqlite_table(
                entry.full_path,
                table_name=table_name_to_load,
                max_rows=self._TABLE_MAX_ROWS,
                max_cols=self._TABLE_MAX_COLS,
            )
        except SQLitePreviewError as e:
            _log.exception("Error reading SQLite database %s", entry.full_path)
            self._render_table_widget(entry.name, [], [], f"Error reading database: {e}")
            return

        def _build_db_ui():
            if len(table.tables) > 1:
                with dpg.group(horizontal=True, parent=self._ctx.panel_id):
                    dpg.add_text("Table:", color=[200, 200, 200])
                    dpg.add_combo(
                        items=table.tables,
                        default_value=table.table_name,
                        width=200,
                        callback=lambda s, a, u: self._render_sqlite_preview(entry, a),
                    )

        self._render_table_widget(entry.name, table.headers, table.rows, table.status, ui_builder=_build_db_ui)
